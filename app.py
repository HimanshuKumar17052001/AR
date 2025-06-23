############################################################
# streamlit_app.py – End-to-end financial‑tables extractor #
############################################################
# 2025‑06‑23 – Revision 3
# • Fix: New company collections were not created because ingest.py's
#   create_collection ran at import‑time with a stale name. We now ensure
#   the (possibly new) collection exists *after* patching the globals and
#   before calling ingest_pdf().
# ---------------------------------------------------------------------------
# Launch with:
#   streamlit run streamlit_app.py
############################################################

from __future__ import annotations

import os
import io
import json
import uuid
import logging
import tempfile
import re
from typing import Dict, List, Tuple, Optional

import streamlit as st
import pandas as pd
import fitz  # PyMuPDF
import requests
from io import StringIO
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from dotenv import load_dotenv
from qdrant_client.http import models as qm

# Re‑use backend modules (must sit in same folder)
import ingest as ingest_mod

# Import ingestion tracker
from ingestion_tracker import (
    is_file_ingested,
    get_ingestion_info,
    extract_company_and_fy_from_pdf_path,
    get_collection_name,
)

# Import configuration
from config import OCR_SERVICE_URL, LOG_LEVEL

# ---------------------------------------------------------------------------
load_dotenv()

# Configure logging
logging.basicConfig(
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    level=getattr(logging, LOG_LEVEL.upper(), logging.INFO),
)
logger = logging.getLogger("streamlit_app")

# ---------------------------------------------------------------------------
FINANCIAL_TABLE_PHRASES: List[str] = [
    "Standalone Balance Sheet",
    "Balance Sheet",
    "Standalone Statement of Balance Sheet",
    "Statement of Balance Sheet",
    "Standalone Profit and Loss Account",
    "Profit and Loss Account",
    "Profit and Loss",
    "Statement of Profit and Loss",
    "Standalone Statement of Profit and Loss",
    "Standalone Cash Flow Statement",
    "Cash Flow Statement",
    "Standalone Statement of Cash Flow",
    "Statement of Cash Flow",
    "Standalone Statement of Cash Flows",
    "Statement of Cash Flows",
    "Consolidated Balance Sheet",
    "Consolidated Profit and Loss Account",
    "Consolidated Statement of Profit and Loss",
    "Consolidated Cash Flow Statement",
    "Consolidated Statement of Cash Flows",
    "Consolidated Statement of Cash Flow",
    "Consolidated Statement of Profit and Loss Account",
]

# ---------------------------------------------------------------------------
# Helper utilities ----------------------------------------------------------
# ---------------------------------------------------------------------------


def is_pre_ingested(fname: str) -> Tuple[Optional[str], Optional[str]]:
    """Check if a file is pre-ingested and return collection name and vector name."""
    if is_file_ingested(fname):
        ingestion_info = get_ingestion_info(fname)
        return ingestion_info["COLLECTION_NAME"], ingestion_info["VECTOR_NAME"]
    else:
        return None, None


# ---------------------------------------------------------------------------
# Search logic functions ----------------------------------------------------
# ---------------------------------------------------------------------------


def normalize_text(text: str) -> str:
    """Normalize text for better keyword matching."""
    # Convert to lowercase and remove extra whitespace
    text = text.lower().strip()
    # Replace multiple spaces with single space
    text = re.sub(r"\s+", " ", text)
    return text


def extract_table_info(summary: str, pdf_filename: str) -> str:
    """
    Extract the appropriate part of the summary based on PDF filename.
    For ICICI_2023-34.pdf: use after first full stop to last
    For all other files: use before first full stop (main heading part)
    """
    # Get just the filename without extension
    filename = os.path.basename(pdf_filename).replace(".pdf", "").replace(".PDF", "")

    # Special case for ICICI_2023-34.pdf
    if filename == "ICICI_2023-34":
        # Find the first full stop
        first_period = summary.find(".")
        if first_period != -1:
            # Return everything after the first full stop (table description part)
            return summary[first_period + 1 :].strip()
        # If no full stop found, return the whole summary
        return summary
    else:
        # For all other files: use before first full stop (main heading part)
        first_period = summary.find(".")
        if first_period != -1:
            # Return everything before the first full stop (main heading part)
            return summary[:first_period].strip()
        # If no full stop found, return the whole summary
        return summary


def phrase_search_in_text(text: str, phrase: str) -> Tuple[bool, int]:
    """
    Check if a phrase exists in the text (case-insensitive).
    Returns (found, position) where position is the index where phrase was found (-1 if not found).
    """
    # Convert both text and phrase to lowercase for case-insensitive search
    normalized_text = normalize_text(text)
    normalized_phrase = normalize_text(phrase)

    position = normalized_text.find(normalized_phrase)
    found = position != -1

    return found, position


def search_phrases_in_collection_with_logic(
    phrases: List[str], collection_name: str, pdf_path: str, qdrant_client
) -> Dict[str, List[Dict[str, Any]]]:
    """
    Search for exact phrases in all summaries in the collection using the special logic.
    Returns results grouped by phrase.
    """
    logger.info(f"Searching for {len(phrases)} phrases in collection {collection_name}")

    # Get all points from collection
    all_points = []
    offset = None
    batch_size = 100

    while True:
        result = qdrant_client.scroll(
            collection_name=collection_name,
            limit=batch_size,
            offset=offset,
            with_payload=True,
            with_vectors=False,  # We don't need vectors for keyword search
        )

        points, next_offset = result
        all_points.extend(points)

        if next_offset is None:
            break
        offset = next_offset

    logger.info(f"Retrieved {len(all_points)} points from collection")

    results = {phrase: [] for phrase in phrases}

    for point in all_points:
        payload = point.payload
        summary = payload.get("summary", "")

        if not summary:
            continue

        # Extract appropriate part based on PDF filename
        search_text = extract_table_info(summary, pdf_path)

        # Check each phrase in the extracted text
        for phrase in phrases:
            found, position = phrase_search_in_text(search_text, phrase)
            if found:
                results[phrase].append(
                    {
                        "page_num": payload.get("page_num"),
                        "position": position,
                        "summary": summary,
                        "search_text": search_text,  # Add the extracted search text
                        "company_name": payload.get("company_name", ""),
                        "phrase": phrase,
                    }
                )

    # Sort each phrase's results by position (earlier in text = higher priority)
    for phrase in results:
        results[phrase].sort(key=lambda x: x["position"])

    return results


# ---------------------------------------------------------------------------
# OCR / Excel helpers -------------------------------------------------------
# ---------------------------------------------------------------------------


def page_to_png(pdf_path: str, page: int, tmp_dir: str, dpi: int = 300) -> str:
    idx = page - 1
    doc = fitz.open(pdf_path)
    if idx >= len(doc):
        raise ValueError(f"Page {page} out of range")
    pix = doc[idx].get_pixmap(matrix=fitz.Matrix(dpi / 72, dpi / 72))
    out = os.path.join(tmp_dir, f"page_{page:04d}.png")
    pix.save(out)
    doc.close()
    return out


def ocr_markdown(img_path: str) -> str:
    with open(img_path, "rb") as f:
        files = {"file": (os.path.basename(img_path), f.read(), "image/png")}
    r = requests.post(
        OCR_SERVICE_URL,
        files=files,
        headers={"accept": "application/json"},
        timeout=300,
    )
    r.raise_for_status()
    jd = r.json()
    if jd.get("status") != "success":
        raise RuntimeError(jd)
    return jd.get("markdown", "")


def md_to_df(md: str) -> Optional[pd.DataFrame]:
    lines = [ln for ln in md.strip().split("\n") if "|" in ln]
    data_lines = [
        ln for ln in lines if not re.fullmatch(r"\|?\s*-+\s*(\|\s*-+\s*)+\|?", ln)
    ]
    if len(data_lines) < 2:
        return None
    df = pd.read_csv(
        StringIO("\n".join(data_lines)), sep="|", engine="python", skipinitialspace=True
    )
    df.columns = df.columns.str.strip()
    return df.dropna(axis=1, how="all")


def add_sheet(wb: Workbook, df: pd.DataFrame, page: int, title: str):
    ws = wb.create_sheet(title=f"Page_{page}")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    for r, row in enumerate(dataframe_to_rows(df, index=False, header=True), 3):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            if r == 3 or c == 1:
                cell.font = Font(bold=True)
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in ws.iter_rows(
        min_row=3, max_row=len(df) + 2, min_col=1, max_col=len(df.columns)
    ):
        for cell in row:
            cell.border = thin
    for col in ws.columns:
        width = min(max(len(str(c.value or "")) for c in col) + 2, 50)
        ws.column_dimensions[col[0].column_letter].width = width


def extract_tables(pdf: str, page_map: Dict[str, List[int]], company: str) -> str:
    page_to_phrase = {p: ph for ph, pages in page_map.items() for p in pages}
    tmp = tempfile.mkdtemp()
    wb = Workbook()
    wb.remove(wb.active)
    for page in sorted(page_to_phrase):
        try:
            img = page_to_png(pdf, page, tmp)
            df = md_to_df(ocr_markdown(img))
            if df is not None and not df.empty:
                add_sheet(wb, df, page, f"{page_to_phrase[page]} – Page {page}")
            else:
                logger.warning("No table on page %s", page)
        except Exception as e:
            logger.error("Page %s error: %s", page, e)
    out = os.path.join(tmp, f"{company}_financial_statement.xlsx")
    wb.save(out)
    return out


# ---------------------------------------------------------------------------
# Streamlit UI -------------------------------------------------------------
# ---------------------------------------------------------------------------

st.set_page_config(
    page_title="Financial Table Extractor",
    layout="centered",
    initial_sidebar_state="expanded",
)
st.title("📊 Financial Statement Table Extractor")

# Add instructions sidebar
with st.sidebar:
    st.header("📋 Instructions")
    st.markdown(
        """
    ### 📁 File Upload Requirements:
    
    **File Format**: PDF only
    
    **Naming Convention**: `COMPANY_FINANCIAL_YEAR.pdf`
    
    **Examples**:
    - `ICICI_2023-24.pdf`
    - `HDFC_2024-25.pdf`
    - `AXIS_2023-24.pdf`
    - `ZOMATO_2024-25.pdf`
    
    **Important**: 
    - Use underscore (_) to separate company name and financial year
    - Company name will be automatically converted to uppercase
    - Financial year should be in format: YYYY-YY or YYYY
    
    ### 📊 Output Format:
    
    **File Name**: `COMPANY_financial_statement.xlsx`
    
    **Content**: Excel workbook with multiple worksheets
    
    **Worksheets Include**:
    - Standalone Balance Sheet
    - Standalone Profit & Loss Account
    - Standalone Cash Flow Statement
    - Consolidated Balance Sheet
    - Consolidated Profit & Loss Account
    - Consolidated Cash Flow Statement
    
    **Formatting**:
    - Bold headers and first column
    - Table borders
    - Auto-sized columns
    - Professional layout
    
    ### 🔄 Process Flow:
    1. Upload your PDF file
    2. System processes the document
    3. Download the Excel file with extracted tables
    """
    )

# Main content area
col1, col2 = st.columns([2, 1])

with col1:
    st.header("📁 Upload Annual Report")

    # File naming rules reminder
    st.info(
        """
    **📝 File Naming Rules:**
    - Format: `COMPANY_FINANCIAL_YEAR.pdf`
    - Example: `ICICI_2023-24.pdf`
    - Use underscore (_) between company name and year
    """
    )

    upload = st.file_uploader(
        "Choose a PDF file",
        type=["pdf"],
        accept_multiple_files=False,
        help="Upload your annual report PDF file with proper naming convention",
    )

with col2:
    st.header("📋 What You'll Get")
    st.markdown(
        """
    **📊 Excel File**: `COMPANY_financial_statement.xlsx`
    
    **📋 Worksheets**:
    - Standalone financial statements
    - Consolidated financial statements
    - Professional formatting
    
    **🎯 Extracted Tables**:
    - Balance Sheet
    - Profit & Loss Account
    - Cash Flow Statement
    """
    )

if upload:
    file_key = f"{upload.name}_{getattr(upload, 'size', 0)}"

    if st.session_state.get("current_file_id") == file_key:
        pass
    else:
        st.session_state.current_file_id = file_key
        st.session_state.excel_bytes = None

        with st.spinner("Processing – please wait..."):
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tf:
                tf.write(upload.read())
                pdf_path = tf.name
            company, _ = extract_company_and_fy_from_pdf_path(upload.name)

            # ------------ Ingestion or reuse ------------
            pre_col, pre_vec = is_pre_ingested(upload.name)
            if pre_col:
                collection = pre_col
                vector = pre_vec
                logger.info("Using pre‑ingested collection %s", collection)
                st.info(
                    f"✅ File {upload.name} is already ingested in collection '{collection}'. Skipping ingestion process."
                )
            else:
                collection = f"{company}_AR_EMBEDDINGS"
                vector = f"{company.lower()}_pagewise_embedding"
                # Patch ingest globals
                ingest_mod.COMPANY_NAME = company
                ingest_mod.COLLECTION_NAME = collection
                ingest_mod.VECTOR_NAME = vector
                # Ensure collection exists *now*
                if not ingest_mod.qdrant_client.collection_exists(collection):
                    sparse_cfg = {
                        vector: qm.SparseVectorParams(index=qm.SparseIndexParams())
                    }
                    ingest_mod.qdrant_client.create_collection(
                        collection, vectors_config={}, sparse_vectors_config=sparse_cfg
                    )
                    logger.info("Created new collection %s", collection)
                # Run ingest
                if not ingest_mod.ingest_pdf(pdf_path):
                    st.error("Ingestion failed (see logs)")
                    st.stop()
                else:
                    st.success(
                        f"✅ Successfully ingested {upload.name} into collection '{collection}'"
                    )

            # ------------ Keyword search ---------------
            # Use the new search function with special logic for ICICI_2023-34.pdf
            result = search_phrases_in_collection_with_logic(
                FINANCIAL_TABLE_PHRASES, collection, pdf_path, ingest_mod.qdrant_client
            )
            page_map = {
                ph: [m["page_num"] for m in matches]
                for ph, matches in result.items()
                if matches
            }
            if not any(page_map.values()):
                st.error("No financial statements detected.")
                st.stop()

            # ------------ Table extraction -------------
            try:
                excel_path = extract_tables(pdf_path, page_map, company)
                with open(excel_path, "rb") as f:
                    st.session_state.excel_bytes = f.read()
            except Exception as e:
                logger.exception("Extraction error: %s", e)
                st.error("Table extraction failed (see logs)")

    if st.session_state.get("excel_bytes"):
        company, _ = extract_company_and_fy_from_pdf_path(upload.name)
        expected_filename = f"{company}_financial_statement.xlsx"

        st.success("✅ Processing Complete!")

        # Show what was processed
        st.info(
            f"""
        **📊 Processed**: {upload.name}
        **🏢 Company**: {company}
        **📁 Output File**: {expected_filename}
        """
        )

        # Download section
        st.header("📥 Download Your Results")

        col3, col4 = st.columns([1, 1])

        with col3:
            st.download_button(
                "📊 Download Excel File",
                st.session_state.excel_bytes,
                file_name=expected_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Download the Excel file with extracted financial tables",
            )

        with col4:
            st.markdown(
                """
            **📋 File Contents**:
            - Multiple worksheets
            - Standalone & Consolidated statements
            - Professional formatting
            - Auto-sized columns
            """
            )

        # Footer note
        st.markdown("---")
        st.markdown(
            """
        **💡 Tip**: The Excel file contains all extracted financial tables with proper formatting. 
        Each worksheet represents a different financial statement from your annual report.
        """
        )
