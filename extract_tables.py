import os
import json
import logging
import requests
import tempfile
from typing import List, Dict, Any, Optional
from pathlib import Path
import fitz  # PyMuPDF
import pandas as pd
from io import StringIO
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from dotenv import load_dotenv

# Import from keyword_search.py
from keyword_search import FINANCIAL_TABLE_PHRASES

# Import ingestion tracker
from ingestion_tracker import (
    is_file_ingested, 
    get_ingestion_info,
    extract_company_and_fy_from_pdf_path,
    get_collection_name
)

# ---------------------------------------------------------------------------
# ENV & CONSTANTS
# ---------------------------------------------------------------------------
load_dotenv()

# Get PDF path from environment or use default
PDF_PATH: str = os.getenv(
    "PDF_PATH", r"C:\Users\himan\Downloads\Documents\ICICI_2023-24.pdf"
)

# Extract company and FY from PDF path
COMPANY_NAME, FINANCIAL_YEAR = extract_company_and_fy_from_pdf_path(PDF_PATH)

# Get collection name - check if file is already ingested first
filename = os.path.basename(PDF_PATH)
if is_file_ingested(filename):
    ingestion_info = get_ingestion_info(filename)
    COLLECTION_NAME = ingestion_info['COLLECTION_NAME']
    COMPANY_NAME = ingestion_info['COMPANY_NAME']
    logger = logging.getLogger("extract_tables")
    logger.info(f"Using existing ingestion info for {filename}: Collection={COLLECTION_NAME}, Company={COMPANY_NAME}")
else:
    # Use default collection naming logic for new files
    COLLECTION_NAME = get_collection_name(PDF_PATH)

OCR_SERVICE_URL: str = "http://52.7.81.94:8000/ocr_image"

# ---------------------------------------------------------------------------
# LOGGING
# ---------------------------------------------------------------------------
LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()
logging.basicConfig(
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    level=getattr(logging, LOG_LEVEL, logging.INFO),
)
logger = logging.getLogger("extract_tables")

# ---------------------------------------------------------------------------
# HELPER FUNCTIONS
# ---------------------------------------------------------------------------


def load_search_results() -> Dict[str, Any]:
    """Load the search results from the keyword search."""
    try:
        with open(f"{COMPANY_NAME}_phrase_search_page_numbers.json", "r") as f:
            return json.load(f)
    except FileNotFoundError:
        logger.error(
            f"{COMPANY_NAME}_phrase_search_page_numbers.json not found. Run keyword_search.py first."
        )
        return {}


def create_folder_structure() -> tuple:
    """Create the folder structure for the company."""
    # Create company folder
    company_folder = COMPANY_NAME
    os.makedirs(company_folder, exist_ok=True)

    # Create downloaded_files and OCR-IMG subfolders
    excel_folder = os.path.join(company_folder, "downloaded_files")
    ocr_img_folder = os.path.join(company_folder, "OCR-IMG")

    os.makedirs(excel_folder, exist_ok=True)
    os.makedirs(ocr_img_folder, exist_ok=True)

    logger.info(f"Created folder structure: {company_folder}/")
    logger.info(f"  - downloaded_files: {excel_folder}")
    logger.info(f"  - OCR-IMG: {ocr_img_folder}")

    return excel_folder, ocr_img_folder


def pdf_page_to_png(
    pdf_path: str, page_num: int, output_path: str, dpi: int = 300
) -> bool:
    """
    Convert a PDF page to PNG image.
    page_num is 1-indexed (as stored in Qdrant)
    """
    try:
        # Open the PDF using the correct PyMuPDF API
        pdf_document = fitz.open(pdf_path)

        # Convert to 0-indexed
        page_index = page_num - 1

        if page_index >= len(pdf_document):
            logger.error(
                f"Page {page_num} does not exist in PDF (max pages: {len(pdf_document)})"
            )
            return False

        # Get the page
        page = pdf_document[page_index]

        # Create transformation matrix for high DPI
        mat = fitz.Matrix(dpi / 72, dpi / 72)

        # Render page to image
        pix = page.get_pixmap(matrix=mat)

        # Save as PNG
        pix.save(output_path)

        pdf_document.close()
        logger.info(f"Converted page {page_num} to PNG: {output_path}")
        return True

    except Exception as e:
        logger.error(f"Failed to convert page {page_num} to PNG: {e}")
        return False


def send_to_ocr_service(image_path: str) -> Optional[Dict[str, Any]]:
    """Send image to OCR service and return the response."""
    try:
        # Read the image file
        with open(image_path, "rb") as f:
            image_data = f.read()

        # Prepare the multipart form data
        files = {"file": (os.path.basename(image_path), image_data, "image/png")}

        # Add headers
        headers = {"accept": "application/json"}

        # Send request
        response = requests.post(
            OCR_SERVICE_URL, files=files, headers=headers, timeout=60
        )

        # Check response
        if response.status_code == 200:
            result = response.json()
            if result.get("status") == "success":
                logger.info(f"OCR successful for {image_path}")
                return result
            else:
                logger.error(f"OCR failed for {image_path}: {result}")
                return None
        else:
            logger.error(
                f"OCR service returned status {response.status_code}: {response.text}"
            )
            return None

    except Exception as e:
        logger.error(f"Failed to send {image_path} to OCR service: {e}")
        return None


def markdown_to_dataframe(markdown: str) -> Optional[pd.DataFrame]:
    """Convert markdown table to pandas DataFrame."""
    try:
        # Split markdown into lines
        lines = markdown.strip().split("\n")

        # Find the table part (lines with |)
        table_lines = [line for line in lines if "|" in line and line.strip()]

        if not table_lines:
            logger.warning("No table lines found in markdown")
            return None

        # Convert markdown table to DataFrame
        # Remove the separator line (contains only ---)
        data_lines = []
        for line in table_lines:
            if not all(cell.strip() == "---" for cell in line.split("|")[1:-1]):
                data_lines.append(line)

        if len(data_lines) < 2:  # Need at least header and one data row
            logger.warning("Insufficient data in markdown table")
            return None

        # Convert to DataFrame using StringIO from io module
        df = pd.read_csv(
            StringIO("\n".join(data_lines)),
            sep="|",
            engine="python",
            skipinitialspace=True,
        )

        # Clean up column names (remove whitespace)
        df.columns = df.columns.str.strip()

        # Remove empty columns
        df = df.dropna(axis=1, how="all")

        return df

    except Exception as e:
        logger.error(f"Failed to convert markdown to DataFrame: {e}")
        return None


def get_filename(
    phrase: str,
    existing_files: List[str],
    page_num: int,
    all_page_phrases: Dict[int, List[str]],
) -> str:
    """Generate filename for the table, handling duplicates with correct naming and consolidated prefix."""
    # Clean phrase for filename
    clean_phrase = phrase.replace(" ", "_").replace("&", "and").replace("/", "_")

    # Check if this page should have "consolidated" prefix
    should_add_consolidated = should_add_consolidated_prefix(page_num, all_page_phrases)

    # If page has both regular and consolidated versions, prefer consolidated
    if should_add_consolidated and not phrase.lower().startswith("consolidated"):
        # Check if there's a consolidated version of this phrase on the same page
        page_phrases = all_page_phrases.get(page_num, [])
        consolidated_versions = [
            p for p in page_phrases if p.lower().startswith("consolidated")
        ]
        if consolidated_versions:
            # Use the consolidated version instead
            consolidated_phrase = consolidated_versions[0]
            clean_phrase = (
                consolidated_phrase.replace(" ", "_")
                .replace("&", "and")
                .replace("/", "_")
            )
        else:
            # Add consolidated prefix
            clean_phrase = f"consolidated_{clean_phrase}"

    base_filename = clean_phrase

    # Check if file already exists
    if base_filename in existing_files:
        # Find the highest _CONTD number
        contd_files = [
            f for f in existing_files if f.startswith(base_filename + "_CONTD")
        ]
        if contd_files:
            # Extract numbers and find max
            numbers = []
            for f in contd_files:
                try:
                    # Handle both _CONTD1 and _CONTD_1 formats
                    if "_CONTD_" in f:
                        num = int(f.split("_CONTD_")[-1])
                    else:
                        num = int(f.split("_CONTD")[-1])
                    numbers.append(num)
                except ValueError:
                    numbers.append(0)
            next_num = max(numbers) + 1
            return f"{base_filename}_CONTD_{next_num}"
        else:
            return f"{base_filename}_CONTD"

    return base_filename


def should_add_consolidated_prefix(
    page_num: int, all_page_phrases: Dict[int, List[str]]
) -> bool:
    """Determine if a page should have 'consolidated' prefix based on its position after Cash Flow Statement."""
    # Get all page numbers sorted
    all_pages = sorted(all_page_phrases.keys())

    # Find the last Cash Flow Statement page
    last_cash_flow_page = None
    for page in all_pages:
        phrases = all_page_phrases[page]
        if any("cash flow statement" in phrase.lower() for phrase in phrases):
            last_cash_flow_page = page

    # If no Cash Flow Statement found, return False
    if last_cash_flow_page is None:
        return False

    # Check if current page comes after the last Cash Flow Statement page
    return page_num > last_cash_flow_page


def save_formatted_excel_worksheet(
    wb, df: pd.DataFrame, page_num: int, phrase: str, is_consolidated: bool = False
):
    """Add a formatted worksheet to the workbook for a specific page."""
    try:
        # Create worksheet name (Excel has 31 character limit for sheet names)
        sheet_name = f"Page_{page_num}"

        # Create new worksheet
        ws = wb.create_sheet(title=sheet_name)

        # Check if phrase already contains standalone or consolidated
        phrase_lower = phrase.lower()
        already_has_prefix = (
            "standalone" in phrase_lower or "consolidated" in phrase_lower
        )

        # Determine the proper title based on phrase and consolidated status
        if already_has_prefix:
            # If phrase already has standalone/consolidated, use it as is
            title = f"{phrase} - Page {page_num}"
        elif is_consolidated:
            if "balance sheet" in phrase.lower():
                title = f"Consolidated Balance Sheet - Page {page_num}"
            elif "profit" in phrase.lower() and "loss" in phrase.lower():
                title = f"Consolidated Profit & Loss Account - Page {page_num}"
            elif "cash flow" in phrase.lower():
                title = f"Consolidated Cash Flow Statement - Page {page_num}"
            else:
                title = f"Consolidated {phrase} - Page {page_num}"
        else:
            if "balance sheet" in phrase.lower():
                title = f"Standalone Balance Sheet - Page {page_num}"
            elif "profit" in phrase.lower() and "loss" in phrase.lower():
                title = f"Standalone Profit & Loss Account - Page {page_num}"
            elif "cash flow" in phrase.lower():
                title = f"Standalone Cash Flow Statement - Page {page_num}"
            else:
                title = f"Standalone {phrase} - Page {page_num}"

        # Add title
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)

        # Write data starting from row 3
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 3):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Make headers bold (row 3)
                if r_idx == 3:
                    cell.font = Font(bold=True)

                # Make first column bold (column A)
                if c_idx == 1:
                    cell.font = Font(bold=True)

        # Add borders around the table
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Apply borders to all cells in the table
        for row in ws.iter_rows(
            min_row=3, max_row=len(df) + 2, min_col=1, max_col=len(df.columns)
        ):
            for cell in row:
                cell.border = thin_border

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width

        logger.info(f"Added worksheet: {sheet_name} - {title}")

    except Exception as e:
        logger.error(f"Failed to add worksheet for page {page_num}: {e}")


def determine_consolidated_status(
    page_num: int,
    all_page_phrases: Dict[int, List[str]],
    current_page_phrases: List[str],
) -> bool:
    """Determine if a page should be labeled as consolidated based on position after standalone Cash Flow Statement."""
    # Get all page numbers sorted
    all_pages = sorted(all_page_phrases.keys())

    # Find the first Consolidated Balance Sheet page
    consolidated_balance_sheet_page = None
    for page in all_pages:
        phrases = all_page_phrases[page]
        if any("consolidated balance sheet" in phrase.lower() for phrase in phrases):
            consolidated_balance_sheet_page = page
            break  # First occurrence of consolidated balance sheet

    # If no Consolidated Balance Sheet found, everything is standalone
    if consolidated_balance_sheet_page is None:
        return False

    # Everything after and including the Consolidated Balance Sheet is consolidated
    return page_num >= consolidated_balance_sheet_page


def extract_tables_from_pages():
    """Main function to extract tables from identified pages."""
    logger.info("Starting table extraction process...")

    # Create folder structure
    excel_folder, ocr_img_folder = create_folder_structure()

    # Load search results
    search_results = load_search_results()
    if not search_results:
        return

    pdf_path = search_results.get("pdf_path")
    page_numbers = search_results.get("page_numbers", {})

    logger.info(f"PDF Path: {pdf_path}")
    logger.info(f"Found {len(page_numbers)} phrases with page numbers")

    # Create a mapping of page numbers to their phrases for consolidated logic
    all_page_phrases = {}
    for phrase, pages in page_numbers.items():
        for page_num in pages:
            if page_num not in all_page_phrases:
                all_page_phrases[page_num] = []
            all_page_phrases[page_num].append(phrase)

    # Get all unique page numbers and sort them in ascending order
    all_pages = sorted(all_page_phrases.keys())
    logger.info(f"Processing {len(all_pages)} pages in order: {all_pages}")

    # Create a single Excel workbook
    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)

    # Process pages in ascending order
    for page_num in all_pages:
        phrases = all_page_phrases[page_num]

        logger.info(f"\n{'='*60}")
        logger.info(f"Processing page {page_num}")
        logger.info(f"Phrases found: {phrases}")
        logger.info(f"{'='*60}")

        # Create PNG filename for OCR-IMG folder
        png_filename = f"page_{page_num:03d}.png"
        png_path = os.path.join(ocr_img_folder, png_filename)

        try:
            # Convert PDF page to PNG and save to OCR-IMG folder
            if not pdf_page_to_png(pdf_path, page_num, png_path):
                continue

            # Send to OCR service
            ocr_result = send_to_ocr_service(png_path)
            if not ocr_result:
                continue

            # Extract markdown
            markdown = ocr_result.get("markdown", "")
            if not markdown:
                logger.warning(f"No markdown returned for page {page_num}")
                continue

            # Convert markdown to DataFrame
            df = markdown_to_dataframe(markdown)
            if df is None or df.empty:
                logger.warning(f"No valid table data for page {page_num}")
                continue

            # Create only ONE worksheet per page, using the first phrase as the title
            # If there are multiple phrases, we'll use the first one for naming
            primary_phrase = phrases[0]  # Use the first phrase found on this page

            # Determine if this should be labeled as consolidated
            is_consolidated = determine_consolidated_status(
                page_num, all_page_phrases, phrases
            )

            logger.info(
                f"Creating worksheet for page {page_num} with primary phrase: {primary_phrase}"
            )
            logger.info(f"Consolidated status: {is_consolidated}")

            # Add worksheet to the workbook
            save_formatted_excel_worksheet(
                wb, df, page_num, primary_phrase, is_consolidated
            )

        except Exception as e:
            logger.error(f"Error processing page {page_num}: {e}")

    # Save the single Excel file
    excel_path = os.path.join(excel_folder, f"{COMPANY_NAME}_financial_statement.xlsx")
    try:
        wb.save(excel_path)
        logger.info(f"Saved single Excel file: {excel_path}")
        logger.info(f"Total worksheets created: {len(wb.sheetnames)}")
        logger.info(f"Worksheets: {wb.sheetnames}")
    except Exception as e:
        logger.error(f"Failed to save Excel file: {e}")

    logger.info(f"\n{'='*60}")
    logger.info("EXTRACTION COMPLETE")
    logger.info(f"{'='*60}")
    logger.info(
        f"Single Excel file saved to: {COMPANY_NAME}/downloaded_files/{COMPANY_NAME}_financial_statement.xlsx"
    )
    logger.info(f"OCR images saved to: {COMPANY_NAME}/OCR-IMG/")

    # List extracted files
    if os.path.exists(excel_folder):
        files = os.listdir(excel_folder)
        logger.info(f"Total files extracted: {len(files)}")
        for file in sorted(files):
            logger.info(f"  - {file}")


def main():
    """Main function."""
    logger.info("Starting table extraction from PDF pages...")

    # Check if PDF exists
    if not os.path.exists(PDF_PATH):
        logger.error(f"PDF file not found: {PDF_PATH}")
        return

    # Check if search results exist
    if not os.path.exists(f"{COMPANY_NAME}_phrase_search_page_numbers.json"):
        logger.error(
            f"{COMPANY_NAME}_phrase_search_page_numbers.json not found. Run keyword_search.py first."
        )
        return

    # Extract tables
    extract_tables_from_pages()


if __name__ == "__main__":
    main()
